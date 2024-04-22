import unittest
import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from fake_useragent import UserAgent
from time import sleep
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook #libreria para trabajar con srchivos .xlsx
from selenium.webdriver.common.by import By




class Tests(unittest.TestCase):
    def setUp(self):
        au = UserAgent()
        opts = Options()
        opts.add_argument(f"user-agent={au}")   
        #con esto hacemos que no muestre el cartel de que el Chrome esta controlado por un software de automatizacion
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])      
        driver_path = ChromeDriverManager().install()
        if not os.path.exists(driver_path):
            driver_path = ChromeDriverManager().install()
        
        # Pasar las opciones al crear la instancia del controlador WebDriver
        self.driver = webdriver.Chrome(service=Service(driver_path), options=opts)
        
    def test_llenar_formulario(self):
        driver = self.driver
        driver.get("D:/ESTUDIOS/automation_selenium_with_python/html/create_account.html")
        sleep(3)
        
        #le decimos el archivo con el que vamos a trabajar y su ruta de acceso
        filesheet="D:/ESTUDIOS/automation_selenium_with_python/static/datos.xlsx"
        
        #de esta manera se abre el archivo con el que vamos a trabajar
        wb = load_workbook(filesheet) 
        
        #con esto nos traemos el nombre de las hojas del archivo
        hojas = wb.sheetnames
        print(f'nombres de las hojas encontradas en el archivo {hojas}') #nos imprime el nombre de las hojas encontradas
        datos_hoja = wb['Hoja2'] #aca le especificamos con que hoja queremos trabajar y se almacena en esta variable todos los datos que se encuntra en ella
        wb.close() #cerramos el archivo
        
        for i in range(1,2): #aca le voy a dar el rango segun cuantas filas de datos tenga.
            username = datos_hoja[f'A{i}'].value
            email = datos_hoja[f'B{i}'].value
            password = datos_hoja[f'C{i}']  #si no le ponemos el .value nos va a traer las coordenadas del elemento y no su informacion
            confirm_password = datos_hoja[f'D{i}'].value
            
            
            #con los if evitamos que envie un None al formulario
            if username is not None:
                driver.find_element(By.ID, "username").send_keys(username)
                print(f"username: {username}")
                sleep(2)
            if email is not None:
                driver.find_element(By.ID, "email").send_keys(email)
                print(f"Email: {email}")
                sleep(2)
            if password is not None:
                # aca le ponemos el .value como al almacenar la informacion en la variable no lo hicisimos
                # asi nos envia la informacion y no las coordenadas
                driver.find_element(By.ID, "password").send_keys(str(password.value)) 
                print(f"password: {password.value}")
                
                sleep(2)
            if confirm_password is not None:
                driver.find_element(By.ID, "confirm_password").send_keys(str(confirm_password))
                print(f"confirm_password: {confirm_password}")
                sleep(2)
            sleep(2)
            driver.find_element(By.ID, "button").click()
            print("se cargo con exito el formulario")
        
        
        
    def tearDown(self):
        self.driver.quit()
        
if __name__ == "__main__":
    unittest.main()